"""
项目管理系统 - 主应用入口
Flask + SQLite + 现代化 Web UI
"""
import os
import json
import shutil
import zipfile
import tempfile
from datetime import datetime, timedelta
from functools import wraps

from flask import (
    Flask, render_template, request, jsonify, redirect,
    url_for, session, send_file, flash
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from config import Config
from database import init_db, get_db, close_connection
from utils.number_generator import generate_project_number, generate_req_id, generate_material_code
from utils.file_utils import (
    ensure_dir, get_project_file_dir, save_uploaded_file,
    save_drawing_file, move_to_history, get_file_size_str
)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY') or os.urandom(24)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB

config = Config()

# ── 初始化 ──────────────────────────────────────────────────────────────
def setup():
    """首次运行初始化"""
    if not config.data_root:
        return False
    ensure_dir(os.path.join(config.data_root, 'project_data'))
    init_db(config.db_path)
    return True


def configure_hosted_runtime():
    """Configure data storage when running on a hosted platform such as Render."""
    data_root = os.environ.get('MIMOCLAW_DATA_ROOT', '').strip()
    if not data_root:
        return False
    config.data_root = os.path.expanduser(data_root)
    setup()
    close_connection(config.db_path)
    return True


configure_hosted_runtime()


# ── 认证装饰器 ──────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json:
                return jsonify({'error': '请先登录'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if session.get('role') not in roles:
                return jsonify({'error': '权限不足'}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator


def log_action(action, target_type='', target_id=0, details=''):
    """记录审计日志"""
    try:
        with get_db(config.db_path) as conn:
            conn.execute(
                "INSERT INTO audit_logs (user_id, action, target_type, target_id, details) VALUES (?,?,?,?,?)",
                (session.get('user_id'), action, target_type, target_id, details)
            )
    except Exception:
        pass


# ── 页面路由 ─────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))


@app.route('/healthz')
def healthz():
    return jsonify({'status': 'ok'})


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username', '')
        password = data.get('password', '')

        if not config.data_root or not os.path.exists(config.db_path):
            if request.is_json:
                return jsonify({'setup_required': True, 'redirect': url_for('setup_page')})
            return redirect(url_for('setup_page'))

        with get_db(config.db_path) as conn:
            user = conn.execute("SELECT * FROM users WHERE username=? AND is_active=1", (username,)).fetchone()

        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['real_name'] = user['real_name']
            log_action('登录系统')
            if request.is_json:
                return jsonify({'success': True, 'redirect': url_for('dashboard')})
            return redirect(url_for('dashboard'))

        if request.is_json:
            return jsonify({'error': '用户名或密码错误'})
        flash('用户名或密码错误')

    return render_template('login.html')


@app.route('/logout')
def logout():
    log_action('退出系统')
    session.clear()
    return redirect(url_for('login'))


@app.route('/setup', methods=['GET', 'POST'])
def setup_page():
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        data_root = data.get('data_root', '').strip()
        if not data_root:
            return jsonify({'error': '请输入数据目录路径'})
        data_root = os.path.expanduser(data_root)
        try:
            os.makedirs(data_root, exist_ok=True)
            config.data_root = data_root
            setup()
            return jsonify({'success': True, 'redirect': url_for('login')})
        except Exception as e:
            return jsonify({'error': f'创建目录失败: {e}'})
    return render_template('setup.html')


@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')


# ── 项目管理 API ─────────────────────────────────────────────────────────
@app.route('/api/projects', methods=['GET'])
@login_required
def api_projects_list():
    with get_db(config.db_path) as conn:
        rows = conn.execute("""
            SELECT p.*, u.real_name as manager_name
            FROM projects p LEFT JOIN users u ON p.manager_id = u.id
            ORDER BY p.created_at DESC
        """).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/projects', methods=['POST'])
@login_required
def api_projects_create():
    data = request.get_json()
    with get_db(config.db_path) as conn:
        # 如果没有提供项目号，自动生成
        project_number = data.get('project_number', '')
        if not project_number:
            project_number = generate_project_number(config, data.get('type', ''), conn)

        try:
            conn.execute("""
                INSERT INTO projects (project_number, project_name, customer, type, manager_id,
                    start_date, end_date, budget, status, description)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (
                project_number, data['project_name'], data.get('customer', ''),
                data.get('type', ''), data.get('manager_id'),
                data.get('start_date', ''), data.get('end_date', ''),
                data.get('budget', 0), '进行中', data.get('description', '')
            ))
            project_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            log_action('创建项目', 'project', project_id, f'项目号: {project_number}')
            return jsonify({'success': True, 'project_id': project_id, 'project_number': project_number})
        except Exception as e:
            return jsonify({'error': str(e)})


@app.route('/api/projects/<int:pid>', methods=['GET'])
@login_required
def api_projects_get(pid):
    with get_db(config.db_path) as conn:
        row = conn.execute("""
            SELECT p.*, u.real_name as manager_name
            FROM projects p LEFT JOIN users u ON p.manager_id = u.id WHERE p.id=?
        """, (pid,)).fetchone()
    if not row:
        return jsonify({'error': '项目不存在'}), 404
    return jsonify(dict(row))


@app.route('/api/projects/<int:pid>', methods=['PUT'])
@login_required
def api_projects_update(pid):
    data = request.get_json()
    fields = []
    values = []
    for key in ['project_name', 'customer', 'type', 'start_date', 'end_date', 'budget', 'status', 'description']:
        if key in data:
            fields.append(f"{key}=?")
            values.append(data[key])
    if 'manager_id' in data:
        fields.append("manager_id=?")
        values.append(data['manager_id'])
    if not fields:
        return jsonify({'error': '没有要更新的字段'})
    fields.append("updated_at=CURRENT_TIMESTAMP")
    values.append(pid)
    with get_db(config.db_path) as conn:
        conn.execute(f"UPDATE projects SET {','.join(fields)} WHERE id=?", values)
        log_action('更新项目', 'project', pid)
    return jsonify({'success': True})


@app.route('/api/projects/<int:pid>', methods=['DELETE'])
@login_required
@role_required('系统管理员', '项目管理员')
def api_projects_delete(pid):
    with get_db(config.db_path) as conn:
        conn.execute("DELETE FROM projects WHERE id=?", (pid,))
        log_action('删除项目', 'project', pid)
    return jsonify({'success': True})


@app.route('/api/projects/generate-number', methods=['POST'])
@login_required
def api_generate_project_number():
    data = request.get_json()
    project_type = data.get('type', '')
    with get_db(config.db_path) as conn:
        number = generate_project_number(config, project_type, conn)
    return jsonify({'project_number': number})


# ── 项目文件 API ─────────────────────────────────────────────────────────
@app.route('/api/projects/<int:pid>/files', methods=['GET'])
@login_required
def api_files_list(pid):
    category = request.args.get('category', '')
    with get_db(config.db_path) as conn:
        sql = "SELECT f.*, u.real_name as uploader_name FROM project_files f LEFT JOIN users u ON f.uploader_id=u.id WHERE f.project_id=?"
        params = [pid]
        if category:
            sql += " AND f.category=?"
            params.append(category)
        sql += " ORDER BY f.upload_time DESC"
        rows = conn.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/projects/<int:pid>/files/upload', methods=['POST'])
@login_required
def api_files_upload(pid):
    if 'files' not in request.files:
        return jsonify({'error': '没有选择文件'})

    with get_db(config.db_path) as conn:
        project = conn.execute("SELECT project_number FROM projects WHERE id=?", (pid,)).fetchone()
        if not project:
            return jsonify({'error': '项目不存在'})

    category = request.form.get('category', '立项文件')
    dest_dir = get_project_file_dir(config.upload_root, project['project_number'], category)
    uploaded = []

    with get_db(config.db_path) as conn:
        for f in request.files.getlist('files'):
            if not f.filename:
                continue
            # 计算新版本号
            row = conn.execute(
                "SELECT MAX(version) as mv FROM project_files WHERE project_id=? AND file_name=?",
                (pid, f.filename)
            ).fetchone()
            version = (row['mv'] or 0) + 1

            saved_name, full_path = save_uploaded_file(f, dest_dir, f.filename, version)
            rel_path = os.path.relpath(full_path, config.upload_root)

            conn.execute("""
                INSERT INTO project_files (project_id, file_name, file_path, version, category, uploader_id)
                VALUES (?,?,?,?,?,?)
            """, (pid, f.filename, rel_path, version, category, session['user_id']))

            # 标记旧版本为非当前
            if version > 1:
                conn.execute(
                    "UPDATE project_files SET is_current=0 WHERE project_id=? AND file_name=? AND version<?",
                    (pid, f.filename, version)
                )

            uploaded.append({'file_name': f.filename, 'version': version})

    log_action('上传文件', 'project', pid, f'文件: {[u["file_name"] for u in uploaded]}')
    return jsonify({'success': True, 'uploaded': uploaded})


@app.route('/api/files/<int:fid>/download')
@login_required
def api_file_download(fid):
    with get_db(config.db_path) as conn:
        row = conn.execute("SELECT * FROM project_files WHERE id=?", (fid,)).fetchone()
    if not row:
        return jsonify({'error': '文件不存在'}), 404
    full_path = os.path.join(config.upload_root, row['file_path'])
    if not os.path.exists(full_path):
        return jsonify({'error': '文件已丢失'}), 404
    return send_file(full_path, as_attachment=True, download_name=row['file_name'])


@app.route('/api/files/<int:fid>/preview')
@login_required
def api_file_preview(fid):
    """文件在线预览（浏览器内嵌查看）"""
    with get_db(config.db_path) as conn:
        row = conn.execute("SELECT * FROM project_files WHERE id=?", (fid,)).fetchone()
    if not row:
        return jsonify({'error': '文件不存在'}), 404
    full_path = os.path.join(config.upload_root, row['file_path'])
    if not os.path.exists(full_path):
        return jsonify({'error': '文件已丢失'}), 404
    # 根据扩展名设置 MIME
    import mimetypes
    mime, _ = mimetypes.guess_type(full_path)
    return send_file(full_path, mimetype=mime, as_attachment=False)


@app.route('/api/files/<int:fid>', methods=['DELETE'])
@login_required
def api_file_delete(fid):
    with get_db(config.db_path) as conn:
        row = conn.execute("SELECT * FROM project_files WHERE id=?", (fid,)).fetchone()
        if row:
            full_path = os.path.join(config.upload_root, row['file_path'])
            if os.path.exists(full_path):
                os.remove(full_path)
            conn.execute("DELETE FROM project_files WHERE id=?", (fid,))
            log_action('删除文件', 'file', fid)
    return jsonify({'success': True})


# ── 用户管理 API ─────────────────────────────────────────────────────────
@app.route('/api/users', methods=['GET'])
@login_required
def api_users_list():
    with get_db(config.db_path) as conn:
        rows = conn.execute("SELECT id, username, role, real_name, email, phone, is_active, created_at FROM users ORDER BY id").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/users', methods=['POST'])
@login_required
@role_required('系统管理员')
def api_users_create():
    data = request.get_json()
    with get_db(config.db_path) as conn:
        try:
            conn.execute(
                "INSERT INTO users (username, password_hash, role, real_name, email, phone) VALUES (?,?,?,?,?,?)",
                (data['username'], generate_password_hash(data['password']),
                 data.get('role', '设计工程师'), data.get('real_name', ''),
                 data.get('email', ''), data.get('phone', ''))
            )
            log_action('创建用户', 'user', 0, f"用户名: {data['username']}")
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'error': str(e)})


@app.route('/api/users/<int:uid>', methods=['PUT'])
@login_required
def api_users_update(uid):
    # 只能修改自己或管理员可修改所有人
    if session['user_id'] != uid and session['role'] != '系统管理员':
        return jsonify({'error': '权限不足'}), 403
    data = request.get_json()
    fields, values = [], []
    for key in ['role', 'real_name', 'email', 'phone', 'is_active']:
        if key in data:
            fields.append(f"{key}=?")
            values.append(data[key])
    if 'password' in data and data['password']:
        fields.append("password_hash=?")
        values.append(generate_password_hash(data['password']))
    if not fields:
        return jsonify({'error': '没有要更新的字段'})
    values.append(uid)
    with get_db(config.db_path) as conn:
        conn.execute(f"UPDATE users SET {','.join(fields)} WHERE id=?", values)
        log_action('更新用户', 'user', uid)
    return jsonify({'success': True})


@app.route('/api/users/<int:uid>', methods=['DELETE'])
@login_required
@role_required('系统管理员')
def api_users_delete(uid):
    if uid == session['user_id']:
        return jsonify({'error': '不能删除当前登录用户'})
    with get_db(config.db_path) as conn:
        conn.execute("UPDATE users SET is_active=0 WHERE id=?", (uid,))
        log_action('禁用用户', 'user', uid)
    return jsonify({'success': True})


# ── 需求管理 API ─────────────────────────────────────────────────────────
@app.route('/api/projects/<int:pid>/requirements', methods=['GET'])
@login_required
def api_requirements_list(pid):
    with get_db(config.db_path) as conn:
        rows = conn.execute("""
            SELECT r.*, u.real_name as creator_name
            FROM requirements r LEFT JOIN users u ON r.created_by=u.id
            WHERE r.project_id=? ORDER BY r.id
        """, (pid,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/projects/<int:pid>/requirements', methods=['POST'])
@login_required
def api_requirements_create(pid):
    data = request.get_json()
    with get_db(config.db_path) as conn:
        req_id = data.get('req_id') or generate_req_id(conn, pid)
        requester_name = data.get('requester_name', session.get('real_name', session.get('username', '')))
        requirement_time = data.get('requirement_time', datetime.now().strftime('%Y-%m-%d'))
        conn.execute("""
            INSERT INTO requirements (project_id, req_id, title, description, priority, status, created_by, parent_req_id, requester_name, requirement_time)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (pid, req_id, data['title'], data.get('description', ''),
              data.get('priority', '中'), '草稿', session['user_id'],
              data.get('parent_req_id'), requester_name, requirement_time))
        rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        log_action('创建需求', 'requirement', rid, f'需求ID: {req_id}, 需求人: {requester_name}')
        return jsonify({'success': True, 'id': rid, 'req_id': req_id})


@app.route('/api/requirements/<int:rid>', methods=['PUT'])
@login_required
def api_requirements_update(rid):
    data = request.get_json()
    fields, values = [], []
    for key in ['title', 'description', 'priority', 'status', 'parent_req_id']:
        if key in data:
            fields.append(f"{key}=?")
            values.append(data[key])
    if not fields:
        return jsonify({'error': '没有要更新的字段'})
    values.append(rid)
    with get_db(config.db_path) as conn:
        conn.execute(f"UPDATE requirements SET {','.join(fields)} WHERE id=?", values)
        log_action('更新需求', 'requirement', rid)
    return jsonify({'success': True})


@app.route('/api/requirements/<int:rid>', methods=['DELETE'])
@login_required
def api_requirements_delete(rid):
    with get_db(config.db_path) as conn:
        req = conn.execute("SELECT * FROM requirements WHERE id=?", (rid,)).fetchone()
        if not req:
            return jsonify({'error': '需求不存在'}), 404
        # 只有创建者本人或管理员可以删除
        if req['created_by'] != session['user_id'] and session.get('role') not in ('系统管理员', '项目管理员'):
            return jsonify({'error': '只能删除自己提出的需求'}), 403
        conn.execute("DELETE FROM requirements WHERE id=?", (rid,))
        log_action('删除需求', 'requirement', rid)
    return jsonify({'success': True})


@app.route('/api/requirements/<int:rid>/changes', methods=['GET'])
@login_required
def api_req_changes_list(rid):
    with get_db(config.db_path) as conn:
        rows = conn.execute("""
            SELECT c.*, u.real_name as applicant_name
            FROM req_changes c LEFT JOIN users u ON c.applicant_id=u.id
            WHERE c.req_id=? ORDER BY c.created_at DESC
        """, (rid,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/requirements/<int:rid>/changes', methods=['POST'])
@login_required
def api_req_changes_create(rid):
    data = request.get_json()
    stages = config.get_approval_stages('requirement')
    with get_db(config.db_path) as conn:
        conn.execute("""
            INSERT INTO req_changes (req_id, change_reason, change_content, impact, applicant_id, approval_stages)
            VALUES (?,?,?,?,?,?)
        """, (rid, data.get('change_reason', ''), data.get('change_content', ''),
              data.get('impact', ''), session['user_id'], json.dumps(stages, ensure_ascii=False)))
        cid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        log_action('提交需求变更', 'req_change', cid)
        return jsonify({'success': True, 'id': cid})


# ── 图纸管理 API ─────────────────────────────────────────────────────────
@app.route('/api/projects/<int:pid>/drawings', methods=['GET'])
@login_required
def api_drawings_list(pid):
    with get_db(config.db_path) as conn:
        rows = conn.execute("""
            SELECT d.*, u.real_name as creator_name
            FROM drawings d LEFT JOIN users u ON d.created_by=u.id
            WHERE d.project_id=? ORDER BY d.id DESC
        """, (pid,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/projects/<int:pid>/drawings', methods=['POST'])
@login_required
def api_drawings_create(pid):
    drawing_number = request.form.get('drawing_number', '')
    drawing_name = request.form.get('drawing_name', '')

    with get_db(config.db_path) as conn:
        # 处理文件上传
        file_path = ''
        version = 1
        if 'file' in request.files and request.files['file'].filename:
            f = request.files['file']
            row = conn.execute(
                "SELECT MAX(version) as mv FROM drawings WHERE project_id=? AND drawing_number=?",
                (pid, drawing_number)
            ).fetchone()
            version = (row['mv'] or 0) + 1
            saved_name, full_path = save_drawing_file(f, config.upload_root, drawing_number, version)
            file_path = os.path.relpath(full_path, config.upload_root)

        conn.execute("""
            INSERT INTO drawings (project_id, drawing_number, drawing_name, version, file_path, status, created_by)
            VALUES (?,?,?,?,?,?,?)
        """, (pid, drawing_number, drawing_name, version, file_path, '草稿', session['user_id']))
        did = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        log_action('创建图纸', 'drawing', did, f'图号: {drawing_number}')
        return jsonify({'success': True, 'id': did})


@app.route('/api/drawings/<int:did>', methods=['PUT'])
@login_required
def api_drawings_update(did):
    data = request.get_json()
    fields, values = [], []
    for key in ['drawing_name', 'status']:
        if key in data:
            fields.append(f"{key}=?")
            values.append(data[key])
    if not fields:
        return jsonify({'error': '没有要更新的字段'})
    values.append(did)
    with get_db(config.db_path) as conn:
        conn.execute(f"UPDATE drawings SET {','.join(fields)} WHERE id=?", values)
    return jsonify({'success': True})


@app.route('/api/drawings/<int:did>/download')
@login_required
def api_drawings_download(did):
    with get_db(config.db_path) as conn:
        row = conn.execute("SELECT * FROM drawings WHERE id=?", (did,)).fetchone()
    if not row or not row['file_path']:
        return jsonify({'error': '文件不存在'}), 404
    full_path = os.path.join(config.upload_root, row['file_path'])
    if not os.path.exists(full_path):
        return jsonify({'error': '文件已丢失'}), 404
    return send_file(full_path, as_attachment=True)


@app.route('/api/drawings/<int:did>/ecos', methods=['GET'])
@login_required
def api_draw_ecos_list(did):
    with get_db(config.db_path) as conn:
        rows = conn.execute("""
            SELECT e.*, u.real_name as applicant_name
            FROM drawing_ecos e LEFT JOIN users u ON e.applicant_id=u.id
            WHERE e.drawing_id=? ORDER BY e.created_at DESC
        """, (did,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/drawings/<int:did>/ecos', methods=['POST'])
@login_required
def api_draw_ecos_create(did):
    data = request.get_json()
    stages = config.get_approval_stages('drawing')
    with get_db(config.db_path) as conn:
        conn.execute("""
            INSERT INTO drawing_ecos (drawing_id, change_description, reason, interchangeability, applicant_id, approval_stages)
            VALUES (?,?,?,?,?,?)
        """, (did, data.get('change_description', ''), data.get('reason', ''),
              data.get('interchangeability', ''), session['user_id'],
              json.dumps(stages, ensure_ascii=False)))
        eid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        log_action('提交图纸ECO', 'drawing_eco', eid)
        return jsonify({'success': True, 'id': eid})


@app.route('/api/drawings/<int:did>/preview')
@login_required
def api_drawings_preview(did):
    """图纸在线预览（PDF内嵌查看，DWG/STP提供下载信息）"""
    with get_db(config.db_path) as conn:
        row = conn.execute("SELECT * FROM drawings WHERE id=?", (did,)).fetchone()
    if not row or not row['file_path']:
        return jsonify({'error': '文件不存在'}), 404
    full_path = os.path.join(config.upload_root, row['file_path'])
    if not os.path.exists(full_path):
        return jsonify({'error': '文件已丢失'}), 404
    import mimetypes
    mime, _ = mimetypes.guess_type(full_path)
    return send_file(full_path, mimetype=mime or 'application/octet-stream', as_attachment=False)


@app.route('/api/drawings/<int:did>/upload-new-version', methods=['POST'])
@login_required
def api_drawings_upload_new(did):
    if 'file' not in request.files:
        return jsonify({'error': '没有选择文件'})
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': '没有选择文件'})

    with get_db(config.db_path) as conn:
        drawing = conn.execute("SELECT * FROM drawings WHERE id=?", (did,)).fetchone()
        if not drawing:
            return jsonify({'error': '图纸不存在'})

        new_version = drawing['version'] + 1
        saved_name, full_path = save_drawing_file(f, config.upload_root, drawing['drawing_number'], new_version)
        file_path = os.path.relpath(full_path, config.upload_root)

        # 旧版本移到历史
        if drawing['file_path']:
            old_full = os.path.join(config.upload_root, drawing['file_path'])
            move_to_history(old_full, config.upload_root)

        conn.execute("UPDATE drawings SET version=?, file_path=?, status='草稿' WHERE id=?",
                      (new_version, file_path, did))
        log_action('上传新版图纸', 'drawing', did, f'版本: V{new_version}')

    return jsonify({'success': True, 'version': new_version})


# ── BOM管理 API ──────────────────────────────────────────────────────────
@app.route('/api/projects/<int:pid>/boms', methods=['GET'])
@login_required
def api_boms_list(pid):
    with get_db(config.db_path) as conn:
        rows = conn.execute("SELECT * FROM boms WHERE project_id=? ORDER BY created_at DESC", (pid,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/projects/<int:pid>/boms', methods=['POST'])
@login_required
def api_boms_create(pid):
    data = request.get_json() or {}
    version = data.get('version', 'V1.0')
    with get_db(config.db_path) as conn:
        conn.execute("INSERT INTO boms (project_id, version, status) VALUES (?,?,?)",
                      (pid, version, '草稿'))
        bid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        log_action('创建BOM', 'bom', bid, f'版本: {version}')
        return jsonify({'success': True, 'id': bid})


@app.route('/api/boms/<int:bid>/items', methods=['GET'])
@login_required
def api_bom_items_list(bid):
    with get_db(config.db_path) as conn:
        rows = conn.execute("""
            SELECT bi.*, d.drawing_number, r.req_id as linked_req_id
            FROM bom_items bi
            LEFT JOIN drawings d ON bi.drawing_id=d.id
            LEFT JOIN requirements r ON bi.req_id=r.id
            WHERE bi.bom_id=? ORDER BY bi.id
        """, (bid,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/boms/<int:bid>/items', methods=['POST'])
@login_required
def api_bom_items_create(bid):
    data = request.get_json()
    with get_db(config.db_path) as conn:
        material_code = data.get('material_code') or generate_material_code(conn)
        conn.execute("""
            INSERT INTO bom_items (bom_id, parent_item_id, material_code, name, spec, unit,
                quantity, position, material, drawing_id, req_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (bid, data.get('parent_item_id'), material_code, data['name'],
              data.get('spec', ''), data.get('unit', '个'), data.get('quantity', 1),
              data.get('position', ''), data.get('material', ''),
              data.get('drawing_id'), data.get('req_id')))
        item_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return jsonify({'success': True, 'id': item_id, 'material_code': material_code})


@app.route('/api/bom-items/<int:item_id>', methods=['PUT'])
@login_required
def api_bom_items_update(item_id):
    data = request.get_json()
    fields, values = [], []
    for key in ['name', 'spec', 'unit', 'quantity', 'position', 'material',
                'drawing_id', 'req_id', 'parent_item_id', 'material_code', 'check_status']:
        if key in data:
            fields.append(f"{key}=?")
            values.append(data[key])
    if not fields:
        return jsonify({'error': '没有要更新的字段'})
    values.append(item_id)
    with get_db(config.db_path) as conn:
        conn.execute(f"UPDATE bom_items SET {','.join(fields)} WHERE id=?", values)
    return jsonify({'success': True})


@app.route('/api/bom-items/<int:item_id>', methods=['DELETE'])
@login_required
def api_bom_items_delete(item_id):
    with get_db(config.db_path) as conn:
        conn.execute("DELETE FROM bom_items WHERE id=?", (item_id,))
    return jsonify({'success': True})


@app.route('/api/boms/<int:bid>/duplicate', methods=['POST'])
@login_required
def api_boms_duplicate(bid):
    """复制BOM创建新版本"""
    with get_db(config.db_path) as conn:
        bom = conn.execute("SELECT * FROM boms WHERE id=?", (bid,)).fetchone()
        if not bom:
            return jsonify({'error': 'BOM不存在'})

        # 新版本号
        ver = bom['version']
        try:
            v_num = float(ver.replace('V', '')) + 0.1
            new_ver = f"V{v_num:.1f}"
        except ValueError:
            new_ver = ver + '.1'

        conn.execute("INSERT INTO boms (project_id, version, status) VALUES (?,?,?)",
                      (bom['project_id'], new_ver, '草稿'))
        new_bid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # 复制物料
        items = conn.execute("SELECT * FROM bom_items WHERE bom_id=?", (bid,)).fetchall()
        id_map = {}
        for item in items:
            conn.execute("""
                INSERT INTO bom_items (bom_id, parent_item_id, material_code, name, spec, unit,
                    quantity, position, material, drawing_id, req_id)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, (new_bid, None, item['material_code'], item['name'], item['spec'],
                  item['unit'], item['quantity'], item['position'], item['material'],
                  item['drawing_id'], item['req_id']))
            new_item_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            id_map[item['id']] = new_item_id

        # 修复父子关系
        for item in items:
            if item['parent_item_id'] and item['parent_item_id'] in id_map:
                conn.execute("UPDATE bom_items SET parent_item_id=? WHERE id=?",
                              (id_map[item['parent_item_id']], id_map[item['id']]))

        log_action('复制BOM', 'bom', new_bid, f'新版本: {new_ver}')
        return jsonify({'success': True, 'id': new_bid, 'version': new_ver})


@app.route('/api/boms/<int:bid>/import', methods=['POST'])
@login_required
def api_boms_import(bid):
    """从CSV/Excel导入BOM物料"""
    if 'file' not in request.files:
        return jsonify({'error': '没有选择文件'})
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': '没有选择文件'})

    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    rows_data = []

    try:
        if ext == 'csv':
            import csv, io
            content = f.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                rows_data.append(row)
        elif ext in ('xlsx', 'xls'):
            from openpyxl import load_workbook
            wb = load_workbook(f, read_only=True)
            ws = wb.active
            headers = [str(cell.value or '').strip() for cell in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = val
                rows_data.append(row_dict)
        else:
            return jsonify({'error': '不支持的文件格式，请上传 CSV 或 Excel 文件'})
    except Exception as e:
        return jsonify({'error': f'文件解析失败: {e}'})

    if not rows_data:
        return jsonify({'error': '文件中没有数据'})

    # 字段映射
    field_map = {
        '物料编码': 'material_code', '编码': 'material_code', 'material_code': 'material_code',
        '名称': 'name', '物料名称': 'name', 'name': 'name',
        '规格': 'spec', '规格型号': 'spec', 'spec': 'spec',
        '单位': 'unit', 'unit': 'unit',
        '数量': 'quantity', 'quantity': 'quantity',
        '位置': 'position', '位号': 'position', 'position': 'position',
        '材质': 'material', '材料': 'material', 'material': 'material',
    }

    imported = 0
    errors = []
    with get_db(config.db_path) as conn:
        for i, row in enumerate(rows_data):
            mapped = {}
            for src_key, dst_key in field_map.items():
                if src_key in row and row[src_key] is not None:
                    mapped[dst_key] = str(row[src_key]).strip()
            name = mapped.get('name', '')
            if not name:
                errors.append(f'第{i+2}行: 缺少物料名称')
                continue
            material_code = mapped.get('material_code') or generate_material_code(conn)
            try:
                quantity = float(mapped.get('quantity', 1) or 1)
            except (ValueError, TypeError):
                quantity = 1
            conn.execute("""
                INSERT INTO bom_items (bom_id, material_code, name, spec, unit, quantity, position, material)
                VALUES (?,?,?,?,?,?,?,?)
            """, (bid, material_code, name,
                  mapped.get('spec', ''), mapped.get('unit', '个'),
                  quantity, mapped.get('position', ''), mapped.get('material', '')))
            imported += 1

    log_action('导入BOM', 'bom', bid, f'导入{imported}条物料')
    result = {'success': True, 'imported': imported}
    if errors:
        result['warnings'] = errors
    return jsonify(result)


@app.route('/api/boms/<int:bid>/export', methods=['GET'])
@login_required
def api_boms_export(bid):
    """导出BOM为Excel"""
    import io
    from openpyxl import Workbook

    with get_db(config.db_path) as conn:
        bom = conn.execute("SELECT * FROM boms WHERE id=?", (bid,)).fetchone()
        items = conn.execute("""
            SELECT bi.*, d.drawing_number, r.req_id as linked_req_id
            FROM bom_items bi
            LEFT JOIN drawings d ON bi.drawing_id=d.id
            LEFT JOIN requirements r ON bi.req_id=r.id
            WHERE bi.bom_id=? ORDER BY bi.id
        """, (bid,)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = f"BOM_{bom['version']}"
    headers = ['物料编码', '名称', '规格', '单位', '数量', '位置', '材质', '关联图纸', '关联需求']
    ws.append(headers)

    for item in items:
        item_dict = dict(item)
        ws.append([
            item_dict['material_code'], item_dict['name'], item_dict['spec'],
            item_dict['unit'], item_dict['quantity'], item_dict['position'],
            item_dict['material'], item_dict.get('drawing_number', ''), item_dict.get('linked_req_id', '')
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"BOM_{bom['version']}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/boms/<int:bid>/check-consistency', methods=['POST'])
@login_required
def api_boms_check(bid):
    """BOM一致性校验：检查关联图纸是否为最新版本"""
    issues = []
    with get_db(config.db_path) as conn:
        items = conn.execute("SELECT * FROM bom_items WHERE bom_id=? AND drawing_id IS NOT NULL", (bid,)).fetchall()
        for item in items:
            drawing = conn.execute("SELECT * FROM drawings WHERE id=?", (item['drawing_id'],)).fetchone()
            if not drawing:
                issues.append({'item_id': item['id'], 'issue': '关联图纸不存在'})
                conn.execute("UPDATE bom_items SET check_status='异常' WHERE id=?", (item['id'],))
            elif drawing['status'] != '已发布':
                issues.append({'item_id': item['id'], 'issue': f'图纸 {drawing["drawing_number"]} 未发布'})
                conn.execute("UPDATE bom_items SET check_status='待核查' WHERE id=?", (item['id'],))
            else:
                conn.execute("UPDATE bom_items SET check_status='正常' WHERE id=?", (item['id'],))
    return jsonify({'issues': issues, 'total_checked': len(items)})


# ── 审批中心 API ─────────────────────────────────────────────────────────
@app.route('/api/approvals/pending', methods=['GET'])
@login_required
def api_approvals_pending():
    """获取当前用户的待审批列表"""
    user_role = session.get('role', '')
    pending = []

    with get_db(config.db_path) as conn:
        # 需求变更待审批
        changes = conn.execute("SELECT * FROM req_changes WHERE status='待审批'").fetchall()
        for c in changes:
            stages = json.loads(c['approval_stages']) if c['approval_stages'] else []
            if user_role in stages or user_role == '系统管理员':
                req = conn.execute("SELECT * FROM requirements WHERE id=?", (c['req_id'],)).fetchone()
                pending.append({
                    'type': 'req_change', 'id': c['id'],
                    'title': f"需求变更: {req['req_id'] if req else '未知'} - {req['title'] if req else ''}",
                    'applicant': c['applicant_id'],
                    'created_at': c['created_at'],
                    'stages': stages
                })

        # 图纸ECO待审批
        ecos = conn.execute("SELECT * FROM drawing_ecos WHERE status='待审批'").fetchall()
        for e in ecos:
            stages = json.loads(e['approval_stages']) if e['approval_stages'] else []
            if user_role in stages or user_role == '系统管理员':
                drawing = conn.execute("SELECT * FROM drawings WHERE id=?", (e['drawing_id'],)).fetchone()
                pending.append({
                    'type': 'drawing_eco', 'id': e['id'],
                    'title': f"图纸ECO: {drawing['drawing_number'] if drawing else '未知'}",
                    'applicant': e['applicant_id'],
                    'created_at': e['created_at'],
                    'stages': stages
                })

    return jsonify(pending)


@app.route('/api/approvals/<target_type>/<int:target_id>', methods=['POST'])
@login_required
def api_approvals_decide(target_type, target_id):
    """审批操作"""
    data = request.get_json()
    decision = data.get('decision', '')  # '通过' or '驳回'
    comment = data.get('comment', '')

    with get_db(config.db_path) as conn:
        # 记录审批
        conn.execute("""
            INSERT INTO approval_records (target_type, target_id, approver_id, decision, comment)
            VALUES (?,?,?,?,?)
        """, (target_type, target_id, session['user_id'], decision, comment))

        # 更新目标状态
        if target_type == 'req_change':
            new_status = '已通过' if decision == '通过' else '已驳回'
            conn.execute("UPDATE req_changes SET status=? WHERE id=?", (new_status, target_id))

            if decision == '通过':
                # 需求版本升级
                change = conn.execute("SELECT * FROM req_changes WHERE id=?", (target_id,)).fetchone()
                if change:
                    req = conn.execute("SELECT * FROM requirements WHERE id=?", (change['req_id'],)).fetchone()
                    if req:
                        # 归档旧版本
                        conn.execute("UPDATE requirements SET status='已归档' WHERE id=?", (req['id'],))
                        # 创建新版本
                        conn.execute("""
                            INSERT INTO requirements (project_id, req_id, title, description, priority,
                                status, created_by, version, parent_req_id)
                            VALUES (?,?,?,?,?,?,?,?,?)
                        """, (req['project_id'], req['req_id'], req['title'],
                              change['change_content'] or req['description'],
                              req['priority'], '已发布', req['created_by'],
                              req['version'] + 1, req['parent_req_id']))

        elif target_type == 'drawing_eco':
            new_status = '已通过' if decision == '通过' else '已驳回'
            conn.execute("UPDATE drawing_ecos SET status=? WHERE id=?", (new_status, target_id))

            if decision == '通过':
                eco = conn.execute("SELECT * FROM drawing_ecos WHERE id=?", (target_id,)).fetchone()
                if eco:
                    conn.execute("UPDATE drawings SET status='已发布' WHERE id=?", (eco['drawing_id'],))
                    # 标记关联BOM物料为待核查
                    conn.execute("""
                        UPDATE bom_items SET check_status='待核查'
                        WHERE drawing_id=?
                    """, (eco['drawing_id'],))

        log_action(f'审批{target_type}', target_type, target_id, f'决定: {decision}')

    return jsonify({'success': True})


@app.route('/api/approvals/history', methods=['GET'])
@login_required
def api_approvals_history():
    with get_db(config.db_path) as conn:
        rows = conn.execute("""
            SELECT a.*, u.real_name as approver_name
            FROM approval_records a LEFT JOIN users u ON a.approver_id=u.id
            ORDER BY a.decided_at DESC LIMIT 100
        """).fetchall()
    return jsonify([dict(r) for r in rows])


# ── 审计日志 API ─────────────────────────────────────────────────────────
@app.route('/api/audit-logs', methods=['GET'])
@login_required
def api_audit_logs():
    limit = request.args.get('limit', 50, type=int)
    with get_db(config.db_path) as conn:
        rows = conn.execute("""
            SELECT a.*, u.real_name as user_name
            FROM audit_logs a LEFT JOIN users u ON a.user_id=u.id
            ORDER BY a.timestamp DESC LIMIT ?
        """, (limit,)).fetchall()
    return jsonify([dict(r) for r in rows])


# ── 仪表盘 API ──────────────────────────────────────────────────────────
@app.route('/api/dashboard/stats', methods=['GET'])
@login_required
def api_dashboard_stats():
    with get_db(config.db_path) as conn:
        projects_total = conn.execute("SELECT COUNT(*) as c FROM projects").fetchone()['c']
        projects_active = conn.execute("SELECT COUNT(*) as c FROM projects WHERE status='进行中'").fetchone()['c']
        reqs_total = conn.execute("SELECT COUNT(*) as c FROM requirements").fetchone()['c']
        drawings_total = conn.execute("SELECT COUNT(*) as c FROM drawings").fetchone()['c']
        boms_total = conn.execute("SELECT COUNT(*) as c FROM boms").fetchone()['c']

        # 待审批数
        user_role = session.get('role', '')
        pending_changes = conn.execute("SELECT COUNT(*) as c FROM req_changes WHERE status='待审批'").fetchone()['c']
        pending_ecos = conn.execute("SELECT COUNT(*) as c FROM drawing_ecos WHERE status='待审批'").fetchone()['c']

        # 最近活动
        recent = conn.execute("""
            SELECT * FROM audit_logs
            WHERE action NOT IN ('登录系统', '退出系统')
            ORDER BY timestamp DESC LIMIT 10
        """).fetchall()

    return jsonify({
        'projects_total': projects_total,
        'projects_active': projects_active,
        'reqs_total': reqs_total,
        'drawings_total': drawings_total,
        'boms_total': boms_total,
        'pending_approvals': pending_changes + pending_ecos,
        'recent_activity': [dict(r) for r in recent]
    })


# ── 系统设置 API ─────────────────────────────────────────────────────────
@app.route('/api/settings/config', methods=['GET'])
@login_required
def api_settings_get():
    return jsonify({
        'data_root': config.data_root,
        'project_number_rule': config.get('general', 'project_number_rule'),
        'project_types': config.get_project_types(),
        'approval_stages': {
            'requirement': config.get_approval_stages('requirement'),
            'drawing': config.get_approval_stages('drawing'),
            'bom': config.get_approval_stages('bom'),
        }
    })


@app.route('/api/settings/config', methods=['PUT'])
@login_required
@role_required('系统管理员')
def api_settings_update():
    data = request.get_json()
    if 'project_number_rule' in data:
        config.set('general', 'project_number_rule', data['project_number_rule'])
    if 'data_root' in data:
        new_root = os.path.expanduser(data['data_root'])
        os.makedirs(new_root, exist_ok=True)
        config.data_root = new_root
        setup()
    if 'approval_stages' in data:
        for module, stages in data['approval_stages'].items():
            config.set('approval', f'{module}_stages', ','.join(stages))
    return jsonify({'success': True})


@app.route('/api/settings/backup', methods=['POST'])
@login_required
@role_required('系统管理员')
def api_settings_backup():
    """备份数据库和附件为zip"""
    if not config.data_root:
        return jsonify({'error': '数据目录未设置'})

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f'backup_{timestamp}.zip'
    backup_path = os.path.join(config.data_root, backup_name)

    try:
        with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            # 备份数据库
            if os.path.exists(config.db_path):
                zf.write(config.db_path, 'database.db')
            # 备份附件目录
            for root, dirs, files in os.walk(config.upload_root):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, config.upload_root)
                    zf.write(file_path, f'files/{arcname}')
        log_action('系统备份', 'system', 0, backup_name)
        return jsonify({'success': True, 'file': backup_name, 'path': backup_path})
    except Exception as e:
        return jsonify({'error': f'备份失败: {e}'})


@app.route('/api/settings/restore', methods=['POST'])
@login_required
@role_required('系统管理员')
def api_settings_restore():
    """从zip恢复"""
    if 'backup' not in request.files:
        return jsonify({'error': '没有选择备份文件'})

    f = request.files['backup']
    tmp_path = os.path.join(tempfile.gettempdir(), f.filename)
    f.save(tmp_path)

    try:
        with zipfile.ZipFile(tmp_path, 'r') as zf:
            zf.extractall(config.data_root)
        init_db(config.db_path)
        log_action('系统恢复', 'system', 0)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': f'恢复失败: {e}'})
    finally:
        os.remove(tmp_path)


# ── 审批流配置 API ───────────────────────────────────────────────────────
@app.route('/api/approval-flows', methods=['GET'])
@login_required
def api_approval_flows_list():
    with get_db(config.db_path) as conn:
        rows = conn.execute("SELECT * FROM approval_flows ORDER BY module, id").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/approval-flows', methods=['POST'])
@login_required
@role_required('系统管理员')
def api_approval_flows_create():
    data = request.get_json()
    with get_db(config.db_path) as conn:
        conn.execute("INSERT INTO approval_flows (module, name, stages) VALUES (?,?,?)",
                      (data['module'], data['name'], json.dumps(data.get('stages', []), ensure_ascii=False)))
    return jsonify({'success': True})


@app.route('/api/approval-flows/<int:fid>', methods=['DELETE'])
@login_required
@role_required('系统管理员')
def api_approval_flows_delete(fid):
    with get_db(config.db_path) as conn:
        conn.execute("DELETE FROM approval_flows WHERE id=?", (fid,))
    return jsonify({'success': True})


# ── 启动 ─────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    setup()
    port = int(os.environ.get('PORT', '5000'))
    debug = os.environ.get('FLASK_DEBUG', '').lower() in ('1', 'true', 'yes')
    app.run(host='0.0.0.0', port=port, debug=debug, use_reloader=False, threaded=True)
